"""Serialización de los KPIs de analítica a Excel y CSV (CB-58 / HST-026).

No calcula nada: toma el payload de `AnalyticsService.get_dashboard_kpis()` y lo
aplana a filas, para que el archivo exportado y la pantalla muestren siempre los
mismos números con los mismos filtros.
"""
import csv
import io
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True)

# (título de la hoja, encabezados, extractor de filas desde el payload)
SHEET_SPECS = [
    (
        'Resumen',
        ['Indicador', 'Valor'],
        lambda d: [
            ['Tasa de conversión (%)', d['conversion_rate']['rate_percentage']],
            ['Leads totales', d['conversion_rate']['total_leads']],
            ['Leads convertidos', d['conversion_rate']['converted_leads']],
            ['Tiempo de respuesta promedio (h)', d['response_time']['avg_hours']],
            ['Tiempo de respuesta mediana (h)', d['response_time']['median_hours']],
            ['Leads sin respuesta', d['response_time']['leads_without_response']],
            ['Leads período actual', d['lead_velocity']['current_period']['count']],
            ['Leads período anterior', d['lead_velocity']['previous_period']['count']],
            ['Crecimiento de leads (%)', d['lead_velocity']['growth_rate_percentage']],
            ['Monto esperado', d['payment_collection']['overall']['expected_amount']],
            ['Monto cobrado', d['payment_collection']['overall']['collected_amount']],
            ['Déficit', d['payment_collection']['overall']['deficit_amount']],
            ['Tasa de cobro (%)', d['payment_collection']['overall']['collection_rate_percentage']],
        ],
    ),
    (
        'Conversión por segmento',
        ['Segmento', 'Leads', 'Convertidos', 'Tasa (%)'],
        lambda d: [
            [r['segment'], r['total_leads'], r['converted_leads'], r['rate_percentage']]
            for r in d['conversion_rate']['by_segment']
        ],
    ),
    (
        'Tiempo de respuesta',
        ['Semana', 'Horas promedio', 'Leads'],
        lambda d: [
            [r['period_start'], r['avg_hours'], r['count']]
            for r in d['response_time']['series']
        ],
    ),
    (
        'Velocidad de leads',
        ['Período', 'Leads'],
        lambda d: [[r['period_start'], r['count']] for r in d['lead_velocity']['series']],
    ),
    (
        'Cobro por programa',
        ['Programa', 'Activo', 'Matrículas activas', 'Esperado', 'Cobrado', 'Déficit',
         'Tasa de cobro (%)', 'Crítico'],
        lambda d: [
            [
                r['program_name'],
                'Sí' if r['is_active'] else 'No',
                r['active_enrollment_count'],
                r['expected_amount'],
                r['collected_amount'],
                r['deficit_amount'],
                r['collection_rate_percentage'],
                'Sí' if r['is_critical'] else 'No',
            ]
            for r in d['payment_collection']['by_program']
        ],
    ),
]


def build_filename(extension: str) -> str:
    return f'analitica-boot-tracker-{date.today().isoformat()}.{extension}'


def _cell(value):
    """Decimal -> float (openpyxl no lo escribe) y None -> '' para no imprimir 'None'."""
    if isinstance(value, Decimal):
        return float(value)
    return '' if value is None else value


def _filter_rows(filters_applied: dict):
    labels = {
        'fecha_desde': 'Desde',
        'fecha_hasta': 'Hasta',
        'segment': 'Segmento',
        'campaign': 'Campaña',
    }
    return [[label, filters_applied.get(key) or 'Todos'] for key, label in labels.items()]


def to_xlsx(data: dict) -> bytes:
    """Un libro con una hoja por bloque de KPIs, más los filtros aplicados."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    filters_sheet = workbook.create_sheet('Filtros')
    filters_sheet.append(['Filtro', 'Valor'])
    filters_sheet['A1'].font = HEADER_FONT
    filters_sheet['B1'].font = HEADER_FONT
    for row in _filter_rows(data['filters_applied']):
        filters_sheet.append(row)

    for title, headers, extract in SHEET_SPECS:
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        for column in range(1, len(headers) + 1):
            sheet.cell(row=1, column=column).font = HEADER_FONT
        for row in extract(data):
            sheet.append([_cell(value) for value in row])
        for column in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 26

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def to_csv(data: dict) -> str:
    """Un solo CSV con las secciones una debajo de otra (Excel no lo soporta multi-hoja)."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    writer.writerow(['Filtros aplicados'])
    writer.writerow(['Filtro', 'Valor'])
    writer.writerows(_filter_rows(data['filters_applied']))

    for title, headers, extract in SHEET_SPECS:
        writer.writerow([])
        writer.writerow([title])
        writer.writerow(headers)
        writer.writerows([_cell(value) for value in row] for row in extract(data))

    return buffer.getvalue()
